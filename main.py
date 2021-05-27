from os import error
import os
import requests
from bs4 import BeautifulSoup
import json
from openpyxl import Workbook
import gspread
from oauth2client.service_account import ServiceAccountCredentials
import os.path


#requests
# 1  = crypto coins
# 2 = tourism coins
# 3 = conv coins
def req(type_req):
    #lists url type coin
    crypto_coins = {
        'aeternity' : 'aeternity-hoje/',
        'ark' : 'ark-hoje/',
        'basic-attention-token' : 'basic-attention-token-hoje/',
        'binance-coin' : 'binance-coin-hoje/',
        'bitcoin-cash' : 'bitcoin-cash-hoje/',
        'bitcoin-diamond' : 'bitcoin-diamond-hoje/',
        'bitcoin-gold' : 'bitcoin-gold-hoje/',
        'bitcoin-sv' : 'bitcoin-sv-hoje/',
        'bitshares' : 'bitshares-hoje/',
        'bytom' : 'bytom-hoje/',
        'cardano' : 'cardano-hoje/',
        'dash' : 'dash/',
        'decred' : 'decred/',
        'dentacoin' : 'dentacoin-hoje/',
        'digibyte' : 'digibyte-hoje/',
        'dogecoin' : 'dogecoin-hoje/',
        'electroneum' : 'electroneum-hoje/',
        'eos' : 'eos-hoje/',
        'ethereum' : 'ethereum/',
        'ethereum-classic' : 'ethereum-classic-hoje/',
        'golem' : 'golem-hoje/',
        'icon' : 'icon-hoje/',
        'iota' : 'iota/',
        'komodo' : 'komodo-hoje/',
        'lisk' : 'lisk/',
        'litecoin' : 'litecoin/',
        'monero' : 'monero/',
        'nano' : 'nano-hoje/',
        'nem' : 'nem/',
        'neo' : 'neo/',
        'omisego' : 'omisego/',
        'ontology' : 'ontology-hoje/',
        'populous' : 'populous-hoje/',
        'pundi-x' : 'pundi-x-hoje/',
        'qtum' : 'qtum-hoje/',
        'reddcoin' : 'reddcoin-hoje/',
        'ripple' : 'ripple-hoje/',
        'salt' : 'salt-hoje/',
        'siacoin' : 'siacoin-hoje/',
        'status' : 'status-hoje/',
        'steem' : 'steem-hoje/',
        'stellar-lumens' : 'stellar-lumens-hoje/',
        'stratis' : 'stratis-hoje/',
        'tether' : 'tether-hoje/',
        'tezos' : 'tezos-hoje/',
        'tron' : 'tron-hoje/',
        'trueusd' : 'trueusd-hoje/',
        'usd-coin' : 'usd-coin-hoje/',
        'vechain' : 'vechain-hoje/',
        'verge' : 'verge-hoje/',
        'wanchain' : 'wanchain-hoje/',
        'waves' : 'waves-hoje/',
        'zcash' : 'zcash/',
        'zilliqa' : 'zilliqa-hoje/',
        '0x' : '0x-hoje/',
    }
    tourism_coins = {
        'dolar-turismo' : 'dolar-turismo/',
        'euro-turismo' : 'euro-turismo/',
    }
    conv_coins = {
        'dolar' : '',
        'euro' : 'euro-hoje/',
        'ouro' : 'ouro-hoje/',
        'dolar-australiano' : 'dolar-australiano-hoje/',
        'dolar-canadense' : 'dolar-canadense-hoje/',
        'dolar-neozelandes' : 'dolar-neozelandes-hoje/',
        'franco-suico' : 'franco-suico-hoje/',
        'iene' : 'iene/',
        'libra' : 'libra-hoje/',
        'novo-sol' : 'novo-sol-hoje/',
        'peso-argentino' : 'peso-argentino/',
        'peso-chileno' : 'peso-chileno/',
        'peso-mexicano' : 'peso-mexicano-hoje/',
        'peso-uruguaio' : 'peso-uruguaio/',
        'rublo-russo' : 'rublo-russo-hoje/',
        'won-sul-coreano' : 'won-sul-coreano-hoje/',
        'yuan' : 'yuan-hoje/',
        
    }
    URL = 'https://dolarhoje.com/'
    result = {}
    consult = {}
    #type request
    if type_req == 1:
        consult = crypto_coins
        pass
    elif type_req == 2:
        consult = tourism_coins
        pass
    elif type_req == 3:
        consult = conv_coins
        pass
    else:
        print("Invalid option!!!!")
        return error
    #scrapping with beatifulSoup4
    for name, coin in consult.items():    
        page = requests.get(URL+coin)
        soup = BeautifulSoup(page.content, 'html.parser')
        results = soup.find(id='nacional')
        result[name] = results['value']
    #return list of coins and respective value in BRL    
    return result
#write a dictionary in a excel(xlsx) file    
def write_excel(coin_dict):
    # if file exist, delete for new consult
    if(os.path.isfile('result.xlsx')):
        os.remove('result.xlsx')
    wb = Workbook()

    ws = wb.active
    ws.cell(row = 1,column=1).value = 'Coin'
    ws.cell(row = 1,column=2).value = 'Value in BRL'
    for id , value in enumerate(coin_dict.items()):
        ws.cell(row = id+2,column=1).value = value[0]
        ws.cell(row = id+2,column=2).value = value[1]
    wb.save(filename="result.xlsx")
#write a dictionary in google sheets    
def google_sheets(coin_dict):
    scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
    #load credentials for google api
    credentials = ServiceAccountCredentials.from_json_keyfile_name('credentials.json', scope)

    gc = gspread.authorize(credentials)

    #get the kay of sheet and insert here   
    wks = gc.open_by_key('KEY')

    #open sheet by name
    wks = gc.open('Test Python')

    worksheet = wks.get_worksheet(0)

    column = 1
    cel = 2
    cel = 2
    worksheet.update_cell(1,1,'Coin')
    worksheet.update_cell(2,1,'Value in BRL')
    for coin, value in coin_dict.items():
        worksheet.update_cell(cel,column,coin)
        column = 2
        worksheet.update_cell(cel,column,value)
        cel = cel + 1
        column = 1
        pass

    pass
#requests
# 1  = crypto coins
# 2 = tourism coins
# 3 = conv coins
#exemple of consult using consult of convetional coins and write in a google sheet
#google_sheets(req(3))
